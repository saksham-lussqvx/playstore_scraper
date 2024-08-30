from playwright.sync_api import sync_playwright
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import pandas
import os
import multiprocessing
import time
import json

# define the base url
base_url = "https://play.google.com/store/apps/details?id="
# open the file containing the list of app ids
file_name = "apk_list_full.csv"
# filename for output file



with open(file_name, "r") as f:
    app_ids = f.read().split("\n")


def fetch_app_info(app_id:str, page) -> dict:
    url = base_url + app_id
    page.goto(url, wait_until="domcontentloaded")
    app_info = {}
    app_info["app_id"] = app_id
    #<div id="error-section" class="uaxL4e">We're sorry, the requested URL was not found on this server.</div>
    # first check if the app is available, if not then return None
    if page.query_selector("#error-section"):
        return None
    # get the html
    html = page.content()
    # parse the html
    soup = BeautifulSoup(html, "html.parser")
    # get the title
    #<h1 class="Fd93Bb F5UCq p5VxAd" itemprop="name">QuickBooks Online Accounting</h1>
    title = soup.find("h1", {"itemprop": "name"}).text
    app_info["title"] = title
    # get the category
    #<span jsname="V67aGc" class="VfPpkd-vQzf8d" aria-hidden="true">Books &amp; Reference</span>
    try:
        #<div class="VfPpkd-LgbsSe VfPpkd-LgbsSe-OWXEXe-INsAgc VfPpkd-LgbsSe-OWXEXe-dgl2Hf Rj2Mlf OLiIxf PDpWxe P62QJc LQeN7 LMoCf" jscontroller="nKuFpb" jsaction="click:cOuCgd; mousedown:UX7yZ; mouseup:lbsD7e; mouseenter:tfO1Yc; mouseleave:JywGue; touchstart:p6p2H; touchmove:FwuNnf; touchend:yfqBxc; touchcancel:JMtRjd; focus:AHmuwe; blur:O22p3e; contextmenu:mg9Pef;mlnRJb:fLiPzd;" data-idom-class="Rj2Mlf OLiIxf PDpWxe P62QJc LQeN7 LMoCf" itemprop="genre"><div class="VfPpkd-Jh9lGc"></div><span jsname="V67aGc" class="VfPpkd-vQzf8d" aria-hidden="true">Books &amp; Reference</span><a class="WpHeLc VfPpkd-mRLv6 VfPpkd-RLmnJb" href="/store/apps/category/BOOKS_AND_REFERENCE" aria-label="Books &amp; Reference" jsname="hSRGPd"></a><div class="VfPpkd-J1Ukfc-LhBDec"></div></div>
        main_div = soup.find("div", {"itemprop": "genre"})
        category = main_div.find_all("span", {"jsname": "V67aGc"})[-1].text
        app_info["category"] = category
    except:
        app_info["category"] = ""
    # get the rating
    #<div class="wVqUob"><div class="ClM7O"><div itemprop="starRating"><div class="TT9eCd" aria-label="Rated 4.7 stars out of five stars">4.7<i class="google-material-icons ERwvGb" aria-hidden="true">star</i></div></div></div><div class="g1rdde">33K reviews</div></div>
    try:
        rating = soup.find("div", {"itemprop": "starRating"}).text
        app_info["rating"] = rating
    except:
        app_info["rating"] = ""
    try:
        # get the review count
        review_count = soup.find("div", {"class": "g1rdde"}).text.split()[0]
        # if there is no number in review count then set it to empty string
        # there could be alphabets in the review count
        x = False
        for i in review_count:
            if i.isdigit():
                x = True
                break
        if not x:
            review_count = ""
        app_info["review_count"] = review_count
    except:
        app_info["review_count"] = ""
    # get the download count
    try:
        #<div class="wVqUob"><div class="ClM7O">1M+</div><div class="g1rdde">Downloads</div></div>
        main_tag = soup.find_all("div", {"class": "wVqUob"})
        download_count = ""
        for tag in main_tag:
            if "Downloads" in tag.text:
                download_count = tag.find("div", {"class": "ClM7O"}).text
                break
        app_info["download_count"] = download_count
    except:
        app_info["download_count"] = ""
    # get the icon link
    try:
        #<div class="l8YSdd"><img src="https://play-lh.googleusercontent.com/VX1WlbpWZDSmPLhsBwJ8H6GYnBQD2MQiAr7FgC3KpLhN8NeVx7t_oD9MxSymGiryDg=s48-rw" srcset="https://play-lh.googleusercontent.com/VX1WlbpWZDSmPLhsBwJ8H6GYnBQD2MQiAr7FgC3KpLhN8NeVx7t_oD9MxSymGiryDg=s96-rw 2x" class="T75of QhHVZd" aria-hidden="true" alt="Icon image" itemprop="image" data-iml="24310473.6"><div class="w7Iutd"><div class="wVqUob"><div class="ClM7O"><div itemprop="starRating"><div class="TT9eCd" aria-label="Rated 3.3 stars out of five stars">3.3<i class="google-material-icons ERwvGb" aria-hidden="true">star</i></div></div></div><div class="g1rdde">13.8K reviews</div></div><div class="wVqUob"><div class="ClM7O">1M+</div><div class="g1rdde">Downloads</div></div><div class="wVqUob"><div class="ClM7O"><img src="https://play-lh.googleusercontent.com/EbEX3AN4FC4pu3lsElAHCiksluOVU8OgkgtWC43-wmm_aHVq2D65FmEM97bPexilUAvlAY5_4ARH8Tb3RxQ=w48-h16-rw" srcset="https://play-lh.googleusercontent.com/EbEX3AN4FC4pu3lsElAHCiksluOVU8OgkgtWC43-wmm_aHVq2D65FmEM97bPexilUAvlAY5_4ARH8Tb3RxQ=w96-h32-rw 2x" class="T75of xGa6dd" alt="Content rating" itemprop="image" data-iml="24310447.800000004"></div><div class="g1rdde"><span itemprop="contentRating"><span>Rated for 3+</span></span><div jsaction="click:CnOdef" class="MKV5ee" role="button" tabindex="0" jscontroller="kJXwXb" aria-label="More info about this content rating"><i class="google-material-icons oUaal" aria-hidden="true">info</i></div></div></div></div></div>
        icon_link = soup.find("img", {"itemprop": "image"}).get("src")
        app_info["icon_link"] = icon_link
    except:
        app_info["icon_link"] = ""
    try:
        # get the developer email
        #<a class="Si6A0c RrSxVb" target="_blank" href="mailto:support+mobile@datacamp.com" aria-label="Support email mailto:support+mobile@datacamp.com will open in a new window or tab."><i class="google-material-icons j25Vu" aria-hidden="true">email</i><div class="pZ8Djf"><div class="xFVDSb">Support email</div><div class="pSEeg">support+mobile@datacamp.com</div></div></a>
        emails = soup.find_all("a", {"class": "Si6A0c RrSxVb", "target": "_blank"})
        developer_email = ""
        for email in emails:
            if "mailto" in email.get("href"):
                developer_email = email.get("href").split(":")[-1]
                break
        app_info["developer_email"] = developer_email
    except:
        app_info["developer_email"] = ""
    # get the short description
    try:
        desc_short = soup.find("meta", {"itemprop": "description"}).get("content")
        app_info["desc_short"] = desc_short.replace("|", " ")
    except:
        app_info["desc_short"] = ""
    try:
        # get the long description
        desc_long = soup.find("div", {"data-g-id": "description"}).text
        app_info["desc_long"] = desc_long.replace("|", " ")
    except:
        app_info["desc_long"] = ""

    # return the app info
    return app_info


def main(app_ids, num):
    while True:
        try:
            #filename = f"final_data_{num}.xlsx"
            json_file = f"final_data_{num}.json"
            # try:
            #     if not os.path.exists(filename):
            #         df = pandas.DataFrame(columns=["app_id","title", "category", "rating", "review_count", "download_count", "icon_link", "developer_email", "desc_short", "desc_long"])
            #         df.to_excel(filename, index=False)
            # except:
            #     pass
            p = sync_playwright().start()
            browser = p.chromium.launch(headless=True)
            context = browser.new_context()
            page = context.new_page()
            for app_id in app_ids:
                processed_app_ids = []
                with open("processed_app_ids.txt", "r") as f:
                    processed_app_ids = f.read().split("\n")
                if app_id in processed_app_ids:
                    continue
                try:
                    app_info = fetch_app_info(app_id, page)
                except:
                    # create a new page
                    try:
                        browser.close()
                    except:
                        pass
                    p = sync_playwright().start()
                    browser = p.chromium.launch(headless=True)
                    context = browser.new_context()
                    page = context.new_page()
                    app_info = fetch_app_info(app_id, page)
                if app_info:
                    # save it in the excel file
                    # df = pandas.read_excel(filename)
                    # df_new = pandas.DataFrame([app_info])
                    # df = pandas.concat([df, df_new], ignore_index=True)
                    # df.to_excel(filename, index=False)
                    # write this data to the json file too
                    with open(json_file, "a", encoding="utf-8") as f:
                        f.write(json.dumps(app_info) + "\n")
                    # use openpyxl to write to the excel file
                    # wb = load_workbook(filename)
                    # ws = wb.active
                    # ws.append([app_info["app_id"], app_info["title"], app_info["category"], app_info["rating"], app_info["review_count"], app_info["download_count"], app_info["icon_link"], app_info["developer_email"], app_info["desc_short"], app_info["desc_long"]])
                    # wb.save(filename)
                    # add the app id to the processed app ids file
                    with open("processed_app_ids.txt", "a") as f:
                        f.write(f"{app_id}\n")
                # if app info is None then just add the app id to the file
                else:
                    # df = pandas.read_excel(filename)
                    # df_new = pandas.DataFrame([{"app_id":app_id,"title": "", "category": "", "rating": "", "review_count": "", "download_count": "", "icon_link": "", "developer_email": "", "desc_short": "", "desc_long": ""}])
                    # df = pandas.concat([df, df_new], ignore_index=True)
                    # df.to_excel(filename, index=False)
                    # add the app id to the processed app ids file
                    # use openpyxl to write to the excel file
                    # write this data to the json file too
                    with open(json_file, "a", encoding="utf-8") as f:
                        f.write(json.dumps({"app_id":app_id,"title": "", "category": "", "rating": "", "review_count": "", "download_count": "", "icon_link": "", "developer_email": "", "desc_short": "", "desc_long": ""}) + "\n")
                    # wb = load_workbook(filename)
                    # ws = wb.active
                    # ws.append([app_id, "", "", "", "", "", "", "", "", ""])
                    # wb.save(filename)
                    with open("processed_app_ids.txt", "a") as f:
                        f.write(f"{app_id}\n")
            browser.close()
            break
        except Exception as e:
            print(e)
            try:
                browser.close()
                del p, browser, context, page
            except:
                pass
            time.sleep(5)
            continue

def split_list(l, n):
    k, m = divmod(len(l), n)
    return (l[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(n))

if __name__ == "__main__":
    # open 5 chrome browsers
    num = 10
    app_ids_split = list(split_list(app_ids, num))
    processes = []
    for i in range(num):
        p = multiprocessing.Process(target=main, args=(app_ids_split[i], i))
        p.start()
        processes.append(p)
        time.sleep(2)
    for p in processes:
        p.join()
    print("All processes are done")